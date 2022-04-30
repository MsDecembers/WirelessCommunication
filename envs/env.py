from itertools import product as product
from itertools import combinations
from math import exp, cos, sin
from os.path import dirname, abspath
import math
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from mpl_toolkits.mplot3d import Axes3D
from numpy import random

import openpyxl
import matlab
import matlab.engine
import tensorflow as tf

tf.disable_v2_behavior()
tf.set_random_seed(1)
np.random.seed(1)
dir_path = dirname(abspath(__file__)) + '/位置文件'
CUE_all_coord = np.load(dir_path + '/cue.npy')
coord_cue_random = np.load(dir_path + '/cue_mix.npy')

'''
初步设想:
这个环境的状态可能以缓存矩阵cache_matrix作为状态
动作为改变cache_matrix
'''


class IrsCompMISOEnv:
    def __init__(
            self,
            bs_num, ue_num, p_max, irs_units_num, antenna_num, fov_patch_num, reflect_max, r_min, BW, open_matlab=False
    ):
        '''

        :param bs_num: 基站个数
        :param ue_num: 用户数量
        :param p_max: 基站最大发射功率
        :param irs_units_num: 智能反射面数量,本模型规定只有一个RIS
        :param antenna_num: 基站天线数量
        :param fov_patch_num: fov子块
        :param mec_storage: mec的存储总量
        :param mec_max_computing_resources: mec的计算资源总量
        :param reflect_max: 反射矩阵最大相移系数
        :param r_min: 每个用户最小传输速率
        :param BW: bandwidth
        '''

        self.cr = 1.5
        self.Kb = 10 ** (-9)
        self.Ub = 10 ** 5
        self.ub = 15
        self.time = 0
        self.mec_storage = generate_storage_mec(bs_num, 160, 250)
        self.mec_max_computing_resources = generate_max_computing_resources_mec(bs_num, 5000, 6500)
        self.fov_sizes = generate_fov_size(fov_patch_num, 1, 1)
        self.epsilon = np.zeros([bs_num, fov_patch_num], np.int)
        self.uefov_table = generate_uefov_table(ue_num)
        self.bsfov_table = generate_bsfov_table(self.epsilon)
        self.omegas = generate_omega_random(bs_num, ue_num, antenna_num)
        self.bs_num = bs_num
        self.ue_num = ue_num
        self.p_max = p_max
        self.irs_units_num = irs_units_num
        self.antenna_num = antenna_num
        self.fov_patch_num = fov_patch_num
        self.action_table = gen_action_space(self.fov_patch_num)
        self.reflect_max = reflect_max
        self.r_min = r_min
        self.rendered_fov_sizes = cal_total_rendered_fov_sizes(self.fov_sizes, self.cr)
        self.total_computing_resources = cal_total_computing_resources(self.fov_sizes, self.Kb, self.Ub, self.ub,
                                                                       self.cr)
        self.BW = BW
        N_0_dbm = -174 + 10 * np.log10(BW)
        self.N_0 = np.power(10, ((N_0_dbm - 30) / 10))

        self.cue_coord = coord_cue_random
        self.ch_space = np.zeros(self.ue_num)
        self.bs_coord = None
        self.gfu_max = 1
        self.engine = 0
        if (open_matlab):
            self.engine = matlab.engine.start_matlab()
        self.action = np.zeros(self.bs_num)
        # self.n_reflect = 5 #将反射矩阵的系数划分成几等级
        self.action_irs = 0
        self.reflect = np.ones((self.irs_units_num, self.irs_units_num))
        # self.action_c_p = np.zeros((self.ue_num+self.antenna_num, self.antenna_num))
        self._coord_set()
        # self._gain_calculate()
        self.G, self.G2 = all_G_gain_cal_MISO_splitI(self.time, self.bs_num, self.ue_num, self.antenna_num,
                                                     self.irs_coord, self.cue_coord,
                                                     self.bs_coord, self.reflect, self.irs_units_num)
        self.states = np.concatenate([np.array(self.G).flatten(), self.epsilon.flatten()], axis=0)
        # self.states=self._gain_contact()+self.ch_add_states()+self.p_add_states()+self.reflect_amp_add_states()
        # self.states = self._gain_contact()
        print("MISO协作缓存环境创建完毕！")

    def _coord_set(self):
        '''
        :return: 根据预先的坐标按照不同数量进行选择
        '''
        # self.bs_coord = np.matrix([[0, 0, 0]]).getA()
        # self.cue_coord = CUE_all_coord[:self.ue_num+self.antenna_num, :]
        self.bs_coord = np.array([[0, 0, 0], [5, 20, 0], [20, 10, 0]])
        # self.cue_coord = coord_cue_random[:self.ue_num, :]
        # a=GBU_all_coord[:self.antenna_num, :]
        # self.cue_coord = np.r_[self.cue_coord,a]
        # self.irs_coord = np.matrix([[31, 6, 0]])
        self.irs_coord = np.array([[6, 12, 0]])

        # 将位置plot出来
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        # matplotlib画图中中文显示会有问题，需要这两行设置默认字体

        plt.xlabel('X')
        plt.ylabel('Y')

        print('基站用户位置分布图')

        colors1 = '#00CED1'  # 点的颜色
        colors2 = '#DC143C'
        colors3 = '#7FFFD4'
        colors4 = '#A52A2A'
        colors5 = '#008000'
        area = np.pi ** 2  # 点面积
        # 画散点图
        plt.scatter(self.bs_coord[:, 0], self.bs_coord[:, 1], s=area * 2, marker='o', c=colors1, alpha=0.4, label='基站')
        plt.scatter(self.irs_coord[:, 0], self.irs_coord[:, 1], s=area * 2, marker='s', c=colors2, alpha=0.4,
                    label='反射面')
        plt.scatter(self.cue_coord[0, :, 0], self.cue_coord[0, :, 1], s=area * 2, marker='v', c=colors3, alpha=0.4,
                    label='CUE用户')
        #             label='D2DR')
        plt.legend()
        plt.show()
        # if self.point>0:
        #     new_coord_lst = []
        #     random_count = self.point*15+1
        #     for g0 in range(self.ue_num):
        #         new_coord_lst.append(coord_cue_random[random_count,g0,:])
        #     for g1 in range(self.antenna_num):
        #         new_coord_lst.append(coord_gbu_random[random_count,g1+self.ue_num,:])
        #     self.cue_coord = np.array(new_coord_lst).reshape(self.ue_num+self.antenna_num,3)
        # else:
        # self._gain_calculate()

    def plot_location(self):
        # 将位置plot出来
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        # matplotlib画图中中文显示会有问题，需要这两行设置默认字体
        plt.xlabel('X')
        plt.ylabel('Y')
        print('位置分布图')
        colors1 = '#00CED1'  # 点的颜色
        colors2 = '#DC143C'
        colors3 = '#7FFFD4'
        colors4 = '#A52A2A'
        colors5 = '#008000'
        area = np.pi ** 2  # 点面积
        # 画散点图
        plt.scatter(self.irs_coord[:, 0], self.irs_coord[:, 1], s=area * 2, marker='s', c=colors2, alpha=0.4,
                    label='反射面')
        plt.scatter(self.cue_coord[:, 0], self.cue_coord[:, 1], s=area, marker='v', c=colors3, alpha=0.4, label='CUE用户')
        plt.legend()
        plt.savefig(dir_path + '/location.png', dpi=300)
        plt.show()

    def _gain_contact(self):
        # 将计算出来的信道增益进行拼接作为state
        a = []
        for ue_num in range(self.ue_num):
            for bs_i in range(self.bs_num):
                for ch_i in range(self.antenna_num):
                    which_bs = self.ch_space[ue_num] / self.antenna_num
                    which_ch = self.ch_space[ue_num] % self.antenna_num
                    if bs_i == which_bs and ch_i == which_ch:
                        a.append(self.G[ue_num])
                    else:
                        a.append(0)
        return a

    def reset(self):
        # 重新设置环境
        # if stat=="all":
        #     self._coord_set()
        self.G, self.G2 = all_G_gain_cal_MISO_splitI(self.time, self.bs_num, self.ue_num, self.antenna_num,
                                                     self.irs_coord, self.cue_coord,
                                                     self.bs_coord, self.reflect, self.irs_units_num)
        self.states = np.concatenate([np.array(self.G).flatten(), self.epsilon.flatten()], axis=0)
        return self.states

    # def user_location_random(self):
    #     #随机生成下一步的位置
    #     limit = bs_dist_limit-100
    #     limit_1 = bs_dist_limit-50
    #     zeros_arr = np.array([0]).reshape(-1,1)
    #     for i in range(self.ue_num):
    #         cx = (-1 + 2*np.random.random())* limit
    #         cy = (-1 + 2*np.random.random())* limit
    #         cxy = np.array([cx,cy]).reshape(1,2)
    #         while np.linalg.norm(cxy, axis=1, keepdims=True) > limit_1:
    #             cx = (-1 + 2*np.random.random())* limit
    #             cy = (-1 + 2*np.random.random())* limit
    #             cxy = np.array([cx,cy]).reshape(1,2)
    #         self.cue_coord[i,:] = np.hstack((cxy,zeros_arr))
    #     print('cue新位置随机成功')

    def cal_reward(self, actions):
        '''
        考虑约束的适应度函数值计算
        :return:
        '''
        # if step != 1:
        #     T_old = T_old*(step-1)
        # else:
        #     T_old = T_old
        bs = []
        value = 0
        ''' 
        先满足所有用户需求的fov都能在MEC上找到缓存，且缓存内容所消耗的计算资源不超过MEC，缓存总大小不超过MEC的存储容量
        动作为每一时刻MEC选择缓存的内容
        首先要验证每个基站选择的缓存行动是否满足约束
        '''

        self.epsilon = np.zeros([self.bs_num, self.fov_patch_num])
        exist_flags = np.zeros(self.fov_patch_num)
        for bs in range(self.bs_num):
            chosen_action = actions[bs]
            caches = self.action_table[chosen_action]
            sum_rendered_size = 0
            sum_computing_resources = 0
            for fov in caches:
                self.epsilon[bs, fov] = 1
                exist_flags[fov] = 1
                sum_rendered_size += self.rendered_fov_sizes[fov]
                sum_computing_resources += self.total_computing_resources[fov]
        if (np.sum(exist_flags) < self.fov_patch_num):
            # print("缓存不满足所有用户需求")
            return -500
        if (sum_rendered_size > self.mec_storage[bs]):
            # print("存储容量超过上限")
            return -200
        if (sum_computing_resources > self.mec_max_computing_resources[bs]):
            # print("计算资源超出MEC上限")
            return -200

        # print(action_all)

        '''在不超过存储容量和计算资源上限的前提下，计算UE上的速率，判断是否满足最小速率'''
        self.bsfov_table = generate_bsfov_table(self.epsilon)
        rates = np.zeros([self.ue_num])
        for ue in range(self.ue_num):
            rate = cal_transmit_rate(self.BW, self.G2, self.omegas, ue, self.uefov_table, self.bsfov_table, self.N_0)
            if (rate < self.r_min):
                print("速率小于最低约束", rate)
                return -100
            rates[ue] = rate

        '''在满足最小速率的情况下，判断BS发射功率是否满足约束，先测试速率部分'''
        total_powers = cal_total_power(static_power=10, pbd=5e-3, epsilon=self.epsilon,
                                       omega=self.omegas, bs_num=self.bs_num, fov_num=self.fov_patch_num,
                                       fov_sizes=self.fov_sizes, Kb=self.Kb, Ub=self.Ub, ub=self.ub, cr=self.cr,
                                       computing_resources=self.total_computing_resources)
        for bs in range(self.bs_num):
            if (total_powers[bs] > self.p_max):
                return -50
        reward = (self.bs_num * self.p_max - np.sum(total_powers)) * 2
        return reward

    def step(self, actions, step):
        # self.action_c_p = action_c_p
        r = self.cal_reward(actions)
        new_coord_lst = []
        self.G, self.G2 = all_G_gain_cal_MISO_splitI(step, self.bs_num, self.ue_num, self.antenna_num, self.irs_coord,
                                                     self.cue_coord,
                                                     self.bs_coord, self.reflect, self.irs_units_num)
        states_ = self.states = np.concatenate([np.array(self.G).flatten(), self.epsilon.flatten()], axis=0)
        return r, states_, self.epsilon

    def action_states(self):
        p = []
        for i in range(self.bs_num):
            p.append(self.action[i])
        return p

    def reflect_amp_add_states(self):
        reflect_amp = []
        for i in range(self.irs_units_num):
            reflect_amp.append(self.reflect[i][i])
        return reflect_amp

    def G_tau__add_states(self):
        g_tau = []
        for i in range(self.ue_num):
            for j in range(self.antenna_num):
                if self.G[i][j] ** 2 >= self.tau:
                    g_tau.append(1)
                else:
                    g_tau.append(0)
        return g_tau


# FOV的分辨率
# 2d的FOV数据大小
FOV_2D = 2 * FLAGS.fOV_2DShape[0] * FLAGS.fOV_2DShape[1]
# 3d的FOV数据大小
FOV_3D = (4 / 3) * FOV_2D

# CPU的处理频率
# F_VR = 3 * 10**9
# F_MEC = 10 * 10**9
#
# f_VR = 15
# f_MEC = 15
#
# k_m = 10**(-9)
# k_v = 10**(-9)
#
# E_MEC = 10**(20)
# E_VR = 10**(15)

# np.random.seed(1)
# BW = 40
N_0_dbm = -174 + 10 * log10(FLAGS.BW)
N_0 = np.power(10, (N_0_dbm - 30) / 10)


# N_0 = 10 ** ((N_0_dbm - 30) / 10)
# N_0 =0.00001
# ue_bs_a = 3
# ue_irs_a = 2.2
# irs_bs_a = 2.2
# ue_bs_a = 3.5
# ue_irs_a = 2.5
# irs_bs_a = 2.5
# gfu_bs_a = 3.5#2.5


def G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect):
    '''
    计算综合信道增益G
    h_cue_bs:用户到基站的信道增益
    h_irs_bs：IRS到基站的信道增益，是个一行K列的矩阵
    h_cue_irs：用户到IRS的信道增益，是一个K行一列的矩阵
    reflect：反射矩阵，是一个K行K列的矩阵
    :return:一个综合信道增益的值
    '''
    # print("h_irs_bs",h_irs_bs)
    # print("reflect",reflect)
    h_irs_bs = np.array(h_irs_bs)
    temp = np.dot(h_irs_bs.T.conjugate(), reflect)
    # print("temp",temp)
    h_cue_irs_bs = np.dot(temp, h_cue_irs)
    G = h_cue_bs + h_cue_irs_bs
    return G


def generate_omega_random(bs_num, ue_num, antenna_num):
    lamb = 1
    omegas = []
    for bs in range(bs_num):
        for ue in range(ue_num):
            omega = np.zeros(shape=[antenna_num], dtype=np.complex)
            for i in range(antenna_num):
                theta = 2 * random.random() * math.pi
                omega[i] = lamb * (math.cos(theta) + math.sin(theta) * 1j)
            omegas.append(omega)
    return omegas


def initial_precoding_matrix(bs_num, ue_num, antenna_num):
    lamb = 1
    omegas = []
    for bs in range(bs_num):
        for ue in range(ue_num):
            omega = np.zeros(shape=[antenna_num], dtype=np.complex)
            for i in range(antenna_num):
                theta = 2 * random.random() * math.pi
                omega[i] = lamb * (math.cos(theta) + math.sin(theta) * 1j)
            omegas.append(omega)
    return omegas


def h_gain_cal(coord_a, coord_b, a, small_fading_style, irs_m):
    '''
    :param coord_a:用户或者基站坐标
    :param coord_b:用户或者基站坐标
    :param a:路径损耗系数
    :param small_fading_style:小尺度衰落
    irs_m:irs元件个数
    :return:增益
    '''
    if small_fading_style == 'Rayleigh':
        small = np.random.normal(0, 1 / 2, 1) + np.random.normal(0, 1 / 2, 1) * 1j
        ad = np.array(np.array(coord_a) - np.array(coord_b)).reshape(1, 3)
        d = np.linalg.norm(ad)
        if d == 0:
            d = 0.000001
        h = np.sqrt(0.001 * d ** (-a)) * small
        return h
    else:
        h = []
        for i in range(irs_m):
            small = np.sqrt(2 / 3) * (
                    exp(0) * (cos(np.random.rand() * 2 * np.pi) + sin(np.random.rand() * 2 * np.pi) * 1j)) \
                    + np.sqrt(1 / 3) * (np.random.normal(0, 1 / 2, 1) + np.random.normal(0, 1 / 2, 1) * 1j)
            # small=1
            ad = np.array(np.array(coord_a) - np.array(coord_b)).reshape(1, 3)
            d = np.linalg.norm(ad)
            if d == 0:
                d = 0.000001
            h.append(np.sqrt(0.001 * d ** (-a)) * small)
        return h


def all_G_gain_cal_MISO(time, bs_num, ue_num, antenna_num, irs_coord, ue_coord, bs_coord, reflect, irs_units_num):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G = []
    count = 0
    for ue in range(ue_num):
        for bs in range(bs_num):
            for antenna in range(antenna_num):
                h_cue_bs = h_gain_cal(ue_coord[time, ue, :], bs_coord[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_units_num)
                h_cue_irs = h_gain_cal(ue_coord[time, ue, :], irs_coord, FLAGS.ue_irs_a, "Racian", irs_units_num)
                h_irs_bs = h_gain_cal(irs_coord, bs_coord[bs], FLAGS.irs_bs_a, "Racian", irs_units_num)
                if irs_units_num != 0:
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0])
                else:
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0])

    return G


def all_G_gain_cal_MISO_splitI(time, bs_num, ue_num, antenna_num, irs_coord, ue_coord, bs_coord, reflect,
                               irs_units_num):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G2 = np.zeros([bs_num, ue_num, antenna_num], dtype=np.complex)
    G = []
    for ue in range(ue_num):
        for bs in range(bs_num):
            for antenna in range(antenna_num):
                h_cue_bs = h_gain_cal(ue_coord[time, ue, :], bs_coord[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_units_num)
                h_cue_irs = h_gain_cal(ue_coord[time, ue, :], irs_coord, FLAGS.ue_irs_a, "Racian", irs_units_num)
                h_irs_bs = h_gain_cal(irs_coord, bs_coord[bs], FLAGS.irs_bs_a, "Racian", irs_units_num)

                if irs_units_num != 0:
                    G2[bs, ue, antenna] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0].real)
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0].imag)
                else:
                    G2[bs, ue, antenna] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0].real)
                    G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0].imag)
    return G, G2


def generate_uefov_table(ue_num):
    '''简化模型，目前认为每个用户请求渲染的fov是一致的'''
    rs = np.zeros(ue_num, dtype=np.int)
    for i in range(ue_num):
        rs[i] = i
    return rs


def generate_bsfov_table(epsilon):
    [bs_num, fov_num] = epsilon.shape
    bsfov_table = []
    for fov in range(fov_num):
        bss = []
        for bs in range(bs_num):
            if (epsilon[bs, fov] == 1):
                bss.append(bs)
        bsfov_table.append(bss)
    return bsfov_table


def cal_sinr(G, omegas, ue, uefov_table, bsfov_table, N0):
    target_fov = uefov_table[ue]
    sum_up = 0
    sum_down = N0
    served_bss = bsfov_table[target_fov]
    fov_num = uefov_table.shape[0]
    for bs in served_bss:
        g = G[bs, target_fov, :]
        omega = np.array(omegas[bs * fov_num + ue]).T
        sum_up += np.power(np.linalg.norm(np.dot(omega, g)), 2)
    for i in range(fov_num):
        if (i != target_fov):
            g = G[bs, i, :]
            bss = bsfov_table[i]
            for bs in bss:
                omega = np.array(omegas[bs * fov_num + i]).T
                sum_down += np.power(np.linalg.norm(np.dot(omega, g)), 2)
    return sum_up / sum_down


def cal_transmit_rate(bw, G, omega, ue, uefov_table, bsfov_table, N0):
    sinr = cal_sinr(G, omega, ue, uefov_table, bsfov_table, N0)
    return bw * np.log2((1 + sinr))


def cal_total_rendered_fov_sizes(fov_sizes, cr):
    rs = np.zeros(len(fov_sizes))
    for fov in fov_sizes:
        rs[fov] = cal_rendered_size(fov_sizes[fov], cr)
    return rs


def cal_total_computing_resources(fov_sizes, Kb, Ub, ub, cr):
    computing_resources = np.zeros(len(fov_sizes))
    for fov in range(len(fov_sizes)):
        computing_resources[fov] = cal_rendered_computing_resources(cal_rendered_size(fov_sizes[fov], cr), Kb, Ub, ub)
    return computing_resources


def cal_transmit_power(epsilon, omega, bs_num, fov_num):
    power = np.zeros(bs_num)
    for bs in range(bs_num):
        p = 0
        for fov in range(fov_num):
            p += np.power(np.linalg.norm(np.dot(epsilon[bs, fov].T, omega[bs * fov_num + fov])), 2)
        power[bs] = p
    return power


def cal_rendered_power(pbd, epsilon, bs_num, fov_num, computing_resources):
    power = np.zeros(bs_num)
    for bs in range(bs_num):
        p = 0
        for fov in range(fov_num):
            p += (pbd * epsilon[bs, fov] * computing_resources[fov])
        power[bs] = p
    return power


def cal_total_power(static_power, pbd, epsilon, omega, bs_num, fov_num, fov_sizes, Kb, Ub, ub, cr):
    computing_resources = cal_total_computing_resources(fov_sizes, Kb, Ub, ub, cr)
    transmit_power = cal_transmit_power(epsilon, omega, bs_num, fov_num)
    rendered_power = cal_rendered_power(pbd, epsilon, bs_num, fov_num, computing_resources)
    return transmit_power + rendered_power + np.ones(bs_num) * static_power


def cal_total_power(static_power, pbd, epsilon, omega, bs_num, fov_num, fov_sizes, Kb, Ub, ub, cr, computing_resources):
    transmit_power = cal_transmit_power(epsilon, omega, bs_num, fov_num)
    rendered_power = cal_rendered_power(pbd, epsilon, bs_num, fov_num, computing_resources)
    return transmit_power + rendered_power + np.ones(bs_num) * static_power


def generate_max_computing_resources_mec(mec_num, min_size, max_size):
    '''
    :param mec_num:
    :param min_size:
    :param max_size:
    :return:
    '''
    rs = np.zeros([mec_num], np.int)
    for i in range(mec_num):
        rs[i] = random.randint(min_size, max_size)
    return rs


def generate_storage_mec(mec_num, min_size, max_size):
    '''
    :param mec_num:
    :param min_size:
    :param max_size:
    :return:
    '''
    rs = np.zeros([mec_num], np.int)
    for i in range(mec_num):
        rs[i] = random.randint(min_size, max_size)
    return rs


def generate_fov_size(fov_num, min_size, max_size):
    '''
    :param fov_num:
    :param min_size:
    :param max_size:
    :return:
    '''
    rs = np.zeros([fov_num], np.int)
    for i in range(fov_num):
        rs[i] = random.randint(min_size, max_size)
    return rs


def cal_rendered_size(fov_size, cr):
    '''
    :param fov_size:
    :param cr: 压缩系数
    :return:
    '''
    rs = 3 * 8 * np.power(fov_size, 2) * 2 / cr
    return rs


def cal_rendered_computing_resources(rendered_fov_size, Kb, Ub, ub):
    '''
    :param rendered_fov_size:
    :param Kb: the architecture coefficient, which is related to the CPU architecture of MEC
    :param Ub: the operating frequency of the MEC
    :param ub: the number of CPU cycles required by MEC to process data per bit during rendering
    :return:
    '''
    rs = rendered_fov_size * Kb * np.power(Ub, 2) * ub
    return rs


'''
计算beamforming中的omega
'''


def omega_cal(h):
    omega = np.zeros_like(h).astype(dtype=np.complex)

    return omega


def calculate_reward(caching_matrix):
    reward = 0
    dic = {}.fromkeys(caching_matrix)
    if (len(dic) == len(caching_matrix)):
        reward = -100
    else:
        reward = -500
    return reward


def dist_calc_x(user_state, x, dist, angle, angle_fix, x_min, x_max, max_speed):
    '''
        Use for calculating the distance of movement in each TS in terms of X-axis
        :param user_state:
        :param x:
        :param dist:
        :param angle:
        :param angle_fix:
        :param x_min:
        :param x_max:
        :param max_speed:
        :return:
        '''
    if user_state == 0:  # random
        new_x = x + dist * math.cos(angle)
    else:
        new_x = x + dist * math.cos(angle) + 4 / 5 * max_speed * math.cos(angle_fix)

    while new_x < x_min or new_x > x_max:
        # print('edge_x')
        new_angle = np.pi + angle
        if user_state == 0:  # random
            new_x = x + dist * math.cos(new_angle)
        else:
            new_x = x + dist * math.cos(new_angle) + 4 / 5 * max_speed * math.cos(angle_fix)
    return new_x


def dist_calc_y(user_state, y, dist, angle, angle_fix, y_min, y_max, max_speed):
    '''
        Use for calculating the distance of movement in each TS in terms of Y-axis
        :param user_state:
        :param x:
        :param dist:
        :param angle:
        :param angle_fix:
        :param x_min:
        :param x_max:
        :param max_speed:
        :return:
        '''
    if user_state == 0:  # random
        new_y = y + dist * math.sin(angle)
    else:
        new_y = y + dist * math.sin(angle) + 4 / 5 * max_speed * math.sin(angle_fix)
    while new_y < y_min or new_y > y_max:
        # print('edge_y')
        new_angle = - angle
        if user_state == 0:  # random
            new_y = y + dist * math.sin(new_angle)
        else:
            new_y = y + dist * math.sin(new_angle) + 4 / 5 * max_speed * math.sin(angle_fix)
    return new_y


def all_G_gain_cal(bs_num, irs_coord, ue_num, coord_a, coord_b, reflect, irs_m):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G = np.zeros([bs_num, ue_num], dtype=np.complex)
    # G = np.zeros(bs*cuenum, dtype="complex")
    count = 0
    for ue in range(ue_num):
        for bs in range(bs_num):
            h_cue_bs = h_gain_cal(coord_a[ue], coord_b[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_m)
            h_cue_irs = h_gain_cal(coord_a[ue], irs_coord, FLAGS.ue_irs_a, "Racian", irs_m)
            h_irs_bs = h_gain_cal(irs_coord, coord_b[bs], FLAGS.irs_bs_a, "Racian", irs_m)
            # test1=(cuenum_i + 1) * (bs_i + 1) * (chnum_i + 1) - 1
            if irs_m != 0:
                G[bs, ue] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
            else:
                G[bs, ue] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
            count += 1
    return G

    # for c in range(cuenum):
    #     if channel_space[c] != 0:
    #         which_bs = int(channel_space[c] / chnum)
    #         which_ch = channel_space[c] % chnum
    #         h_cue_bs = h_gain_cal(coord_a[c], coord_b[which_bs], gfu_bs_a, "Rayleigh", irs_m)
    #         h_cue_irs = h_gain_cal(coord_a[c], irs_coord, ue_irs_a, "Racian", irs_m)
    #         h_irs_bs = h_gain_cal(irs_coord, coord_b[which_bs], irs_bs_a, "Racian", irs_m)
    #         if irs_m != 0:
    #             G[c] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
    #         else:
    #             G[c] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
    #     else:
    #         G[c] = 0
    # return G


def sic_decode_judge(cuenum, channel_num, G):
    '''
    判断解码顺序是否满足条件
    '''
    for i in range(channel_num):
        for j in range(cuenum):
            if pow(abs(G[cuenum + i][i]), 2) > pow(abs(G[j][i]), 2) and G[j][i] != 0:
                return -1
    return 1


def r_min_judge(cue, ch_k, G, action_c_ch, action_c_p, r_min):
    for i in range(cue):
        if sum(action_c_p[i, :]) != 1:
            return -1, None
    for i in range(ch_k):
        if sum(action_c_p[:, i]) > 1:
            return -1, None
    r_arr = []
    for i in range(cue):
        for j in range(ch_k):
            if action_c_ch[i][j] == 1:
                r = FLAGS.BW * log2(1 + (pow(abs(G[i][j]), 2) * action_c_p[i, j] / (N_0)))
                r_arr.append(r)
    for i in range(cue):
        if r_arr[i] < r_min:
            return -1, None
    return 1, r_arr

    # for i in range(cue):
    #     r_i=0
    #     for j in range(ch_k):
    #         for k in range(ch_k):
    #             if G[cue+k,j]!=0:
    #                 G_gbu_bs=G[cue+k,j]
    #                 temp=cue+k
    #                 break
    #         i_gbu=action_c_p[temp,j]*pow(abs(G_gbu_bs),2)
    #         i_gfu=0
    #         for k in range(cue):
    #             i_gfu=i_gfu+action_x[j][k][i]*pow(abs(G[k,j]),2)*action_c_p[k,j]
    #         r_i_channel=BW*log2(1+(pow(abs(G[i,j]),2)*action_c_p[i,j]/(i_gfu+i_gbu+N_0)))
    #         r_i=r_i+r_i_channel
    #     r_arr.append(r_i)
    # for i in range(ch_k):
    #     for j in range(ch_k):
    #         if i==j:
    #             r_i=BW*log2(1+(pow(abs(G[cue+i,j]),2)*action_c_p[cue+i,j]/(N_0)))
    #             r_arr.append(r_i)
    # for i in range(cue):
    #     if r_arr[i] < r_min:
    #         return -1
    # return 1


def clean_G(G, cuenum, tau, channel_num):
    for i in range(cuenum):
        for j in range(channel_num):
            if pow(abs(G[i, j]), 2) < tau:
                G[i, j] = 0
    # for j in range(channel_num):
    #     for i in range(cuenum):
    #         if pow(abs(G[i,j]),2)<pow(abs(G[cuenum+j,j]),2):
    #             G[i,j]=0
    return G


def tau_judge_fun(G, channel_num, cuenum, tau):
    for cue in range(cuenum):
        if sum(G[cue, :]) == 0:
            return -1
    return 1


# def fun(cue, ch_k, r_arr, action_c_p):
#     '''
#     计算r
#     '''
#     t_relay = 0
#     for i in range(cue):
#         for j in range(ch_k):
#             if action_c_p[i][j] != 0:
#                 t_relay += C/(cr*r_arr[i])
#     return t_relay
#     # G_gbu_bs=0
#     # G=np.array(G).reshape(cue+ch_k,ch_k)
#     # action_c_p=np.array(action_c_p).reshape(cue+ch_k,ch_k)
#     # temp=0
#     # r=0
#     # throughput=0
#     # r_arr=[]
#     # for i in range(cue):
#     #     r_i=0
#     #     for j in range(ch_k):
#     #         for k in range(ch_k):
#     #             if G[cue+k,j]!=0:
#     #                 G_gbu_bs=G[cue+k,j]
#     #                 temp=cue+k
#     #                 break
#     #         i_gbu=action_c_p[temp,j]*pow(abs(G_gbu_bs),2)
#     #         i_gfu=0
#     #         for k in range(cue):
#     #             i_gfu=i_gfu+action_x[j][k][i]*pow(abs(G[k,j]),2)*action_c_p[k,j]
#     #         r_i_channel=BW*log2(1+(pow(abs(G[i,j]),2)*action_c_p[i,j]/(i_gfu+i_gbu+N_0)))
#     #         r_i=r_i+r_i_channel
#     #     throughput=throughput+r_i
#     #     r_arr.append(r_i)
#     #     # r_i=r_i-r_min
#     #     r=r+r_i
#     # # print(throughput,r)
#     # # print("fun_r_arr",r_arr)
#     # return throughput,r,r_arr

def render_time(loc):
    time = 0
    if loc == 0:
        time = (FOV_2D * FLAGS.f_vr) / (FLAGS.F_vr)
    else:
        time = (FOV_2D * FLAGS.f_mec) / (FLAGS.F_mec)
    # 时间单位换算成毫秒
    return time * 10 ** 3


def qos_judge(G, r_min, cuenum, channel_num, action_c_ch, action_c_p):
    r_min_judge_va, r_arr = r_min_judge(cuenum, channel_num, G, action_c_ch, action_c_p, r_min)
    # sic_decode_result =sic_decode_judge(cuenum, channel_num, G)
    # if r_min_judge_va==-1 or tau_judge==-1 or sic_decode_result == -1 :
    if r_min_judge_va == -1:
        print("r_min_judge_va", r_min_judge_va)
        return -1, None
    return 1, r_arr


def channel_space_generate(ch_k):
    '''
    可用的各种信道矩阵
    '''
    available_space_channel = []
    a = range(2)  # a可以看作列表[0,1]
    for item in product(a,
                        repeat=ch_k):  # product(range(2), repeat=3) --> 000 001 010 011 100 101 110 111,笛卡尔积product(A,repeat=3)等价于product(A,A,A)，product(A, B) 和 ((x,y) for x in A for y in B)一样.
        available_space_channel.append(item)
    return available_space_channel


def reflect_calculate(reflect_action_arr, reflect_amp_arr, irs_m):
    '''
    :param reflect_action_arr: 反射矩阵角度
    :return: 返回计算后的反射矩阵
    '''
    reflect = np.zeros((irs_m, irs_m), dtype=np.complex)
    for i in range(irs_m):
        reflect[i, i] = reflect_amp_arr[i] * exp(0) * (cos(reflect_action_arr[i]) + sin(reflect_action_arr[i]) * 1j)
    return reflect


def x_generate(G, ch_k, cue):
    '''
    可用的各种反射矩阵
    '''
    G = np.array(G).reshape(cue, ch_k)
    action_x = []
    # action_x=np.zeros((cue+ch_k,cue+ch_k,ch_k))
    for i in range(ch_k):
        action_x_ch = np.zeros((cue, cue + ch_k))
        temp_arr = np.array(G[:, i])
        for j in range(cue):
            if temp_arr[j] == 0:
                action_x_ch[j, :] = 0
            for k in range(cue):
                if temp_arr[j] != 0 and pow(abs(temp_arr[j]), 2) <= pow(abs(temp_arr[k]), 2):
                    action_x_ch[j][k] = 1
                if temp_arr[j] != 0 and temp_arr[k] == 0:
                    action_x_ch[j][k] = 1
                if j == k:
                    action_x_ch[j][k] = 0
        action_x.append(action_x_ch)
    return action_x


def plot_mode_irs_compare(number, ave_throughput_arr1, throughput_arr1, in_ave_throughput_arr1, stat_lst):
    plt.figure()
    for stat in range(len(stat_lst)):
        in_ave_throughput_arr = in_ave_throughput_arr1[stat]
        if stat == 0:
            plt.plot(np.arange(0, len(in_ave_throughput_arr), 9), in_ave_throughput_arr[::9], c='#00CED1', marker='*',
                     alpha=0.4, label='C=3')  # np.arange函数返回一个有终点和起点的固定步长的排列
        if stat == 1:
            plt.plot(np.arange(0, len(in_ave_throughput_arr), 9), in_ave_throughput_arr[::9], c='#9932CC', marker='<',
                     alpha=0.4, label='C=2')  # np.arange函数返回一个有终点和起点的固定步长的排列
        if stat == 2:
            plt.plot(np.arange(0, len(in_ave_throughput_arr), 9), in_ave_throughput_arr[::9], c='g', marker='>',
                     alpha=0.4, label='C=1')
        if stat == 3:
            plt.plot(np.arange(0, len(in_ave_throughput_arr), 9), in_ave_throughput_arr[::9], c='#DC143C', marker='o',
                     alpha=0.4, label='NO_IRS')
    plt.grid(linestyle='-.')
    plt.ylabel('interval_ave_Throughput')
    plt.xlabel('Steps')
    dir_path = dirname(abspath(__file__))
    plt.legend(loc='best')
    plt.savefig(dir_path + '/convergence.pdf')  # , dpi=300)
    plt.show()
    plt.figure()
    for stat in range(len(stat_lst)):
        ave_throughput_arr = ave_throughput_arr1[stat]
        if stat == 0:
            plt.plot(np.arange(10000, len(ave_throughput_arr), 9999), ave_throughput_arr[10000:210000:9999],
                     c='#00CED1', marker='*', alpha=0.4, label='C=3')  # np.arange函数返回一个有终点和起点的固定步长的排列
        if stat == 1:
            plt.plot(np.arange(10000, len(ave_throughput_arr), 9999), ave_throughput_arr[10000:210000:9999],
                     c='#9932CC', marker='<', alpha=0.4, label='C=2')  # np.arange函数返回一个有终点和起点的固定步长的排列
        if stat == 2:
            plt.plot(np.arange(10000, len(ave_throughput_arr), 9999), ave_throughput_arr[10000:210000:9999], c='g',
                     marker='>', alpha=0.4, label='C=1')
        if stat == 3:
            plt.plot(np.arange(10000, len(ave_throughput_arr), 9999), ave_throughput_arr[10000:210000:9999],
                     c='#DC143C', marker='o', alpha=0.4, label='NO_IRS')
    plt.grid(linestyle='-.')
    plt.ylabel('Ave_Throughput')
    plt.xlabel('Steps')
    dir_path = dirname(abspath(__file__))
    plt.legend(loc='best')
    plt.savefig(dir_path + '/convergence1.pdf')  # , dpi=300)
    plt.show()
    plt.close()


def npyload(filename):
    """
    :功能：读取npy文件
    :param filename: 文件名称
    :return:
    """
    print('read file: %s' % (filename))
    return np.load(filename, allow_pickle=True).item()


def npysave(data, filename):
    """
    :功能：保存npy文件
    :param data: 数据
    :param filename: 文件名
    :return:
    """
    namearr = re.split(r'[\\/]', filename)

    #   判断操作系统
    sys = platform.system()
    if sys == "Windows":
        pathstr = '\\'.join(namearr[:-1])
    elif sys == "Linux":
        pathstr = '/'.join(namearr[:-1])
    filestr = namearr[-1]
    if not os.path.exists(pathstr):
        print('make dir：%s' % (pathstr))
        os.makedirs(pathstr)
    print('write to: %s' % (filename))
    np.save(filename, data)


def excel_save(excel, irs_m, stat):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("sheet1")
    i = 1
    for a in excel:
        for j in range(irs_m):
            ws.cell(row=i, column=j + 1).value = a[0][j]
        i += 1
    wb.save('data' + stat + '.xlsx')


def ch_max_cue_judge(action_c_p, cuenum, ch_k, gfu_max):
    judge = 0
    # print("p_max",p_max)
    for i in range(ch_k):
        count = 0
        for j in range(cuenum):
            if action_c_p[j][i] != 0:
                count += 1
        if count > gfu_max:
            judge = 1
            break
    if judge == 1:
        return -1
    else:
        return 1


def cue_max_ch_judge(action_c_p, cuenum, ch_k, ch_max):
    judge = 0
    # print("p_max",p_max)
    for i in range(cuenum):
        count = 0
        for j in range(ch_k):
            if action_c_p[i][j] != 0:
                count += 1
        if count > ch_max:
            judge = 1
            break
    if judge == 1:
        return -1
    else:
        return 1


def average_power(ue, bs, ch, ch_space, p_max):
    power_result = np.zeros(ue)
    bs_power = np.zeros((ue, 2))
    for i in range(ue):
        which_bs = int(ch_space[i] / ch)
        bs_power[ue][0] = ue
        bs_power[ue][1] = which_bs
    # for j in range(bs):
    k = 0
    for i in range(ch_space.shape[0]):
        for j in range(ch_space.shape[1]):
            if ch_space[i][j] != 0:
                k += 1
    average_p = p_max / k
    for i in range(ue):
        for j in range(ch):
            if ch_space[i][j] != 0:
                power_result[i][j] = average_p
    return power_result


def gen_action_space(num):
    rs = []

    a = range(num)  # a可以看作列表[0,1]
    for i in range(1, num + 1):
        comb = list(combinations(a, i))
        for set in comb:
            rs.append(set)

    return rs
